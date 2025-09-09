import openpyxl
from openpyxl.styles import PatternFill

source_file = "Financial Sample.xlsx"
target_file = "Financial Sample 2.xlsx"
output_file = "Comparison_File_1.xlsx"

wb_source = openpyxl.load_workbook(source_file)
wb_target = openpyxl.load_workbook(target_file)
wb_output = openpyxl.load_workbook(source_file)  

ws_source = wb_source.active
ws_target = wb_target.active
ws_output = wb_output.active

total_records_source = ws_source.max_row
total_records_target = ws_target.max_row
discrepancy_count = 0

fill_style = PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid")

for row in ws_source.iter_rows():
    for cell in row:
        cell_location = cell.coordinate
        source_value = cell.value
        target_value = ws_target[cell_location].value
        ws_output[cell_location].value = source_value
        if source_value != target_value:
            ws_output[cell_location].fill = fill_style
            discrepancy_count += 1 


wb_output.save(output_file)


print(f"Total records in source file: {total_records_source}")
print(f"Total records in target file: {total_records_target}")
print(f"Total discrepancies found: {discrepancy_count}")
print(f"Comparison file saved as {output_file}")
